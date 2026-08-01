"""
rl_env/export_ppo_eval_excel.py
==================================
train_ppo.py가 만든 rl_env/outputs/ppo_val_eval.csv를 엑셀에서 바로 훑어보기
좋은 .xlsx로 변환한다 (헤더 고정, 실제 포디움 차량 강조, 과도한 피트횟수 강조).

실행:
    python -m rl_env.export_ppo_eval_excel
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent / "outputs"
CSV_PATH = OUT_DIR / "ppo_val_eval.csv"
XLSX_PATH = OUT_DIR / "ppo_val_eval.xlsx"

# 실제 포디움 성공 차량 강조(초록)
PODIUM_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
# 피트횟수가 규칙 기반 평균(약 8회)의 2배를 훌쩍 넘는, 리워드 해킹 의심 구간 강조(빨강)
EXCESSIVE_PIT_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
EXCESSIVE_PIT_COUNT = 20

HEADER_FONT = Font(bold=True)


def export(csv_path: Path = CSV_PATH, xlsx_path: Path = XLSX_PATH) -> Path:
    df = pd.read_csv(csv_path)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ppo_val_eval", index=False)
        ws = writer.sheets["ppo_val_eval"]

        ws.freeze_panes = "A2"  # 헤더 고정
        for cell in ws[1]:
            cell.font = HEADER_FONT

        for row_idx, (podium, pit_count) in enumerate(
            zip(df["REAL_PODIUM"], df["SIM_PIT_COUNT"]), start=2
        ):
            if pit_count >= EXCESSIVE_PIT_COUNT:
                fill = EXCESSIVE_PIT_FILL
            elif podium:
                fill = PODIUM_FILL
            else:
                fill = None
            if fill is not None:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, col_name in enumerate(df.columns, start=1):
            width = max(len(col_name), df[col_name].astype(str).str.len().max()) + 2
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    return xlsx_path


if __name__ == "__main__":
    path = export()
    print(f"엑셀 저장 완료: {path}")
