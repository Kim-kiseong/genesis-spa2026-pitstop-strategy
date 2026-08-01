"""
rl_env/export_baseline_excel.py
==================================
rule_based_strategy.py가 만든 rl_env/outputs/rule_based_baseline.csv를
엑셀에서 바로 훑어보기 좋은 .xlsx로 변환한다 (헤더 고정, 피트인 랩 강조).

실행:
    python -m rl_env.export_baseline_excel
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent / "outputs"
CSV_PATH = OUT_DIR / "rule_based_baseline.csv"
XLSX_PATH = OUT_DIR / "rule_based_baseline.xlsx"

PIT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FONT = Font(bold=True)


def export(csv_path: Path = CSV_PATH, xlsx_path: Path = XLSX_PATH) -> Path:
    df = pd.read_csv(csv_path)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="rule_based_baseline", index=False)
        ws = writer.sheets["rule_based_baseline"]

        ws.freeze_panes = "A2"  # 헤더 고정
        for cell in ws[1]:
            cell.font = HEADER_FONT

        action_col = df.columns.get_loc("ACTION") + 1  # openpyxl은 1-based
        for row_idx, action in enumerate(df["ACTION"], start=2):
            if action == 1:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PIT_FILL

        for col_idx, col_name in enumerate(df.columns, start=1):
            width = max(len(col_name), df[col_name].astype(str).str.len().max()) + 2
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    return xlsx_path


if __name__ == "__main__":
    path = export()
    print(f"엑셀 저장 완료: {path}")
